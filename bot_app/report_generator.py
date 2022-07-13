import datetime
from openpyxl import load_workbook
import os

try:
    from bot_app.SQL_db_model import (
        City,
        Admin,
        Manager,
        PropertyObjects,
        SalePropertyAttributes,
        RentPropertyAttributes,
        LandLords,
        ObjectPhotos,
        Clients,
        ObjectShares,
        AgencyNumber,
    )
except ModuleNotFoundError:
    from SQL_db_model import (
        City,
        Admin,
        Manager,
        PropertyObjects,
        SalePropertyAttributes,
        RentPropertyAttributes,
        LandLords,
        ObjectPhotos,
        Clients,
        ObjectShares,
        AgencyNumber,
    )


class ReportGenerator:
    def __init__(self, bot, db_repo):
        self.bot = bot
        self.db_repo = db_repo

    def generate_report(self, message):
        pass
        cwd = os.getcwd()
        fn_temp = "template.xlsx"
        path = os.path.join(cwd, "temp", fn_temp)
        path_save = os.path.join(cwd, "temp", "report.xlsx")

        wb = load_workbook(path)
        ws = wb["Object Shares"]

        ws.append(["date", datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")])
        day1 = datetime.datetime.now() - datetime.timedelta(days=7)
        day1 = day1.replace(hour=0, minute=0, second=0, microsecond=0)

        ws.append(["Client Name", "Property ID", "Manager Name", "Time Of Adding"])

        shared_objects = (
            ObjectShares.select()
            .where(ObjectShares.time_of_adding > day1)
            .order_by(ObjectShares.time_of_adding)
        )

        for share in shared_objects:
            ws.append(
                [
                    Clients.get(id=share.related_to_client_id).name,
                    PropertyObjects.get(id=share.related_to_property_id).fake_id,
                    Manager.get(id=share.related_to_manager_id).name,
                    share.time_of_adding.strftime("%m/%d/%Y, %H:%M:%S"),
                ]
            )

        ws = wb.create_sheet("Added Objects")

        for manager in self.db_repo.get_all_active_managers_admin():
            ws.append(["Manager:", manager.name])
            ws.append(
                ["Object ID", "Address", "City", "Service  Type", "Time Of Adding"]
            )
            objest_to_show = (
                PropertyObjects.select()
                .where(PropertyObjects.related_to_manager_id == manager)
                .order_by(PropertyObjects.time_of_adding)
            )
            within_week = [
                obj
                for obj in objest_to_show
                if obj.time_of_adding
                > datetime.datetime.now() - datetime.timedelta(days=7)
            ]
            for obj in within_week:
                ws.append(
                    [
                        obj.fake_id,
                        obj.address,
                        City.get(id=obj.related_to_city_id).city_name,
                        obj.property_service_type,
                        obj.time_of_adding.strftime("%m/%d/%Y, %H:%M:%S"),
                    ]
                )
            ws.append([])
            ws.append([])

        wb.save(path_save)
        wb.close()

        with open(path_save, "rb") as file_to_send:
            self.bot.send_document(message.chat.id, file_to_send)
